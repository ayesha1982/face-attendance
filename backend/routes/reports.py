import io
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, date

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/summary', methods=['GET'])
def report_summary():
    from models.employee import Employee, Attendance
    sd = request.args.get('start_date', date.today().replace(day=1).isoformat())
    ed = request.args.get('end_date', date.today().isoformat())
    q  = Attendance.query.filter(
        Attendance.date >= datetime.strptime(sd,'%Y-%m-%d').date(),
        Attendance.date <= datetime.strptime(ed,'%Y-%m-%d').date())
    records = q.all()
    emps    = Employee.query.filter_by(is_active=True).all()
    by_dept = {}
    for r in records:
        dept = r.employee.department or 'Other'
        by_dept.setdefault(dept, {'present':0,'late':0,'absent':0})
        key = r.status if r.status in ['present','late','absent'] else 'present'
        by_dept[dept][key] += 1
    return jsonify({
        'total_employees': len(emps),
        'total_records':   len(records),
        'present': len([r for r in records if r.status=='present']),
        'late':    len([r for r in records if r.status=='late']),
        'absent':  len([r for r in records if r.status=='absent']),
        'by_department': by_dept,
        'period': {'start': sd, 'end': ed}
    })

@reports_bp.route('/excel', methods=['GET'])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500
    from models.employee import Employee, Attendance
    sd = request.args.get('start_date', date.today().replace(day=1).isoformat())
    ed = request.args.get('end_date',   date.today().isoformat())
    q  = Attendance.query.join(Employee)
    if sd: q = q.filter(Attendance.date >= datetime.strptime(sd,'%Y-%m-%d').date())
    if ed: q = q.filter(Attendance.date <= datetime.strptime(ed,'%Y-%m-%d').date())
    records = q.order_by(Attendance.date.desc()).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Attendance'
    hfill = PatternFill(start_color='1C2333', end_color='1C2333', fill_type='solid')
    hfont = Font(name='Calibri', bold=True, color='58A6FF', size=11)
    hdrs  = ['Emp ID','Name','Department','Designation','Date','Check In','Check Out','Status']
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfill and hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
    for row,r in enumerate(records,2):
        ws.append([
            r.employee.emp_id, r.employee.name,
            r.employee.department or '', r.employee.designation or '',
            r.date.strftime('%Y-%m-%d'),
            r.check_in.strftime('%H:%M:%S')  if r.check_in  else '-',
            r.check_out.strftime('%H:%M:%S') if r.check_out else '-',
            r.status.upper()
        ])
    widths = [10,20,18,20,12,12,12,12]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'attendance_{sd}_to_{ed}.xlsx')

@reports_bp.route('/pdf', methods=['GET'])
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        return jsonify({'error': 'reportlab not installed'}), 500
    from models.employee import Employee, Attendance
    sd = request.args.get('start_date', date.today().replace(day=1).isoformat())
    ed = request.args.get('end_date',   date.today().isoformat())
    q  = Attendance.query.join(Employee)
    if sd: q = q.filter(Attendance.date >= datetime.strptime(sd,'%Y-%m-%d').date())
    if ed: q = q.filter(Attendance.date <= datetime.strptime(ed,'%Y-%m-%d').date())
    records = q.order_by(Attendance.date.desc()).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
          rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=16,
                  textColor=colors.HexColor('#58A6FF'), spaceAfter=8)
    elements = [
        Paragraph(f'Attendance Report: {sd} to {ed}', title_style),
        Spacer(1, 0.3*cm)
    ]
    tdata = [['Emp ID','Name','Department','Date','Check In','Check Out','Status']]
    for r in records:
        tdata.append([
            r.employee.emp_id, r.employee.name, r.employee.department or '-',
            r.date.strftime('%d %b %Y'),
            r.check_in.strftime('%H:%M')  if r.check_in  else '-',
            r.check_out.strftime('%H:%M') if r.check_out else '-',
            r.status.upper()
        ])
    col_w = [2.5*cm, 4.5*cm, 4*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    t = Table(tdata, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1C2333')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.HexColor('#58A6FF')),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#30363D')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#0D1117'),colors.HexColor('#161B22')]),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#C9D1D9')),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),('TOPPADDING',(0,0),(-1,-1),5),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=f'attendance_{sd}_to_{ed}.pdf')
